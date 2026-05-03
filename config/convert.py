#!/usr/bin/env python3
"""
Excel → JSON config converter for Godot projects.

Usage:
  python3 convert.py                  # convert all .xlsx in templates/ → json/
  python3 convert.py --template       # generate blank template .xlsx

Excel format:
  - One sheet per entity type (enemy, weapon, player)
  - Row 1 = comment/description row (skipped)
  - Row 2 = property names (first column "id", second "name")
  - Row 3+ = data rows, empty rows skipped
  - "id" = unique key within sheet, "name" = display name

JSON output:
  { "enemy": { "basic": { "name": "Basic Enemy", "hp": 1, ... } } }
"""

import json, os, sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
TEMPLATE_DIR = HERE / "templates"
JSON_DIR = HERE / "json"

TEMPLATE_DATA = {
    "enemy": [
        ["说明", "hp=生命 atk=攻击 chaseSpeed=追踪速度 attackCooldown=攻击间隔(接触伤害冷却) expValue=经验值 isBoss=是否Boss | 攻击范围=CollisionShape2D尺寸，无attackRange字段"],
        ["id", "name", "hp", "maxHp", "atk", "chaseSpeed", "attackCooldown", "expValue", "isBoss"],
        ["wildMan", "野人", 1, 1, 2, 120, 1.0, 1, False],
    ],
    "weapon": [
        ["说明", "atk=伤害 orbitRadius=轨道半径 orbitSpeed=转速 knockbackStrength=击退力"],
        ["id", "name", "atk", "orbitRadius", "orbitSpeed", "knockbackStrength"],
        ["sword", "剑", 1, 80, 4.5, 220],
        ["longSword", "长剑", 3, 100, 2.5, 800],
        ["axe", "斧", 4, 85, 2, 400],
        ["testKnockback", "击退测试武器", 0, 100, 5, 2000],
        ["throwKnife", "飞刀", 1, 75, 6, 50],
    ],
    "player": [
        ["说明", "hp=生命 atk=攻击(武器提供) moveSpeed=移速 weaponSpinSpeed=武器转速 | 开局写入ValueHub，对局内增益系统可动态修改"],
        ["id", "name", "hp", "maxHp", "atk", "moveSpeed", "weaponSpinSpeed"],
        ["default", "剑客", 30, 30, 0, 350, 4.5],
    ],
    "levelSet": [
        ["说明", "expRequired=升到下一级所需经验（独立，非累计）"],
        ["id", "name", "expRequired"],
        ["1", "Lv.1", 10],
        ["2", "Lv.2", 30],
        ["3", "Lv.3", 50],
        ["4", "Lv.4", 100],
        ["5", "Lv.5", 300],
    ],
    "gift": [
        ["说明", "id=技能标识 name=名称 desc=描述(%s替换数值) effectVar=作用变量 values=各级数值(逗号分隔) type=add/*%/magnet"],
        ["id", "name", "desc", "effectVar", "values", "type"],
        ["mighty", "强大", "可以额外增加%s生命值", "maxHp", "10,30,50,80,100", "*%"],
        ["proficiency", "熟练度", "武器攻击速度增加%s", "spinSpeed", "3,5,8,10,20", "*%"],
        ["weaponPlus", "武器+1", "武器数量+1", "weaponCount", "1,1,1,1,1,1,1,1,1,1", "add"],
        ["magnet", "吸铁石", "脉冲吸取%s范围内全部经验", "magnet", "250:8,280:7,320:6,370:5,430:4,500:3,580:2,700:1", "magnet"],
        ["enlarge", "倍大化", "武器变大%s", "weaponScale", "5,10,15,20,25,30,40,50,60", "*%"],
        ["damage", "增伤", "武器伤害增加%s", "weaponAtk", "50,100", "*%"],
    ],
}


def generate_template():
    try:
        import openpyxl
    except ImportError:
        print("pip install openpyxl  required for template generation")
        return

    path = TEMPLATE_DIR / "game_config.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in TEMPLATE_DATA.items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)

    wb.save(path)
    print(f"Template written → {path}")


def convert():
    try:
        import openpyxl
    except ImportError:
        print("pip install openpyxl  required")
        return

    JSON_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_files = list(TEMPLATE_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print("No .xlsx files found in templates/")
        return

    for xlsx_path in xlsx_files:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        config = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 3:
                continue

            # Row 1 = comment (skip), Row 2 = headers, Row 3+ = data
            headers = [str(h).strip() for h in rows[1]]
            if headers[0].lower() != "id":
                print(f"  skip sheet '{sheet_name}': first column must be 'id'")
                continue

            sheet_data = {}
            for row in rows[2:]:
                if row[0] is None:
                    continue
                entry = {}
                for i, h in enumerate(headers):
                    if i >= len(row):
                        break
                    val = row[i]
                    entry[h] = val
                sheet_data[str(row[0])] = entry

            config[sheet_name] = sheet_data

        out_name = xlsx_path.stem + ".json"
        out_path = JSON_DIR / out_name
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2, ensure_ascii=False, default=str)
        print(f"Converted {xlsx_path.name} → {out_path.name}  ({len(config)} sheets)")


if __name__ == "__main__":
    if "--template" in sys.argv:
        generate_template()
    else:
        convert()
